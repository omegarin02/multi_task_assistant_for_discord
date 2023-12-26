import datetime 
import discord
from discord.ext import commands
import tempfile

import os
import openai
import json
from PIL import Image, ImageFilter,  ImageDraw
import shutil
#from openpyxl import Workbook
import openpyxl


from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive

import time

class LiveScheduer:
    def __init__(self,config,gdrive_setting_path=""):
        print("->",gdrive_setting_path)
        gauth = GoogleAuth(gdrive_setting_path) #if gdrive_setting_path != "" else GoogleAuth()
        gauth.LocalWebserverAuth()
        self.drive = GoogleDrive(gauth)
        self.config = config
        self.tmp_dir_path = self.config["files"]["tmp_dir_path"]
        self.icons = self.config["files"]["icons"]
        self.contents=self.config["files"]["contents"]
        self.platform = self.config["files"]["platform"]
        self.schedule = self.config["files"]["schedule"]
        self.number_img = self.config["files"]["number_img"]
        self.column_size = self.config["grid"]["column_size"]
        self.column_first_point = self.config["grid"]["column_first_point"]

    def centering(self,img,column_key):
        x,y = img.size
        column_x = self.column_size[column_key]["x"]
        column_y = self.column_size[column_key]["y"]
        return int ((column_x-x)/2),int ((column_y-y)/2)

    async def print_schedule(self,ctx=None):
        schedule_obj = self.get_schedule()
        base_image = self.get_schedule_base_image()
        liver_images = self.get_liver_icon()
        platform_images = self.get_live_platform_icon()
        contents_images = self.get_live_contents_icon()
        number_picture = self.get_number_picture()
        #draw = ImageDraw.Draw(base_image)
        i = 0
        for key,val in schedule_obj.items():
            for column_key,column_point in self.column_first_point.items():
                loop = 1 if column_key != "liver" else len(val[column_key].keys())
                size = self.column_size[column_key]
                """
                for j in range(loop):
                    draw.line(( column_point["x"],
                                column_point["y"]+size["y"]*i,
                                column_point["x"]+size["x"]*(j+1),
                                column_point["y"]+size["y"]*i ),
                                fill=(255, 0, 255), width=1)
                    draw.line(( column_point["x"]+size["x"]*(j+1),
                                column_point["y"]+size["y"]*i,
                                column_point["x"]+size["x"]*(j+1),
                                column_point["y"]+size["y"]*(i+1)),
                                fill=(255, 0, 255), width=1)
                """
                if (column_key == "day"):
                    j = 0
                    for d in key:
                        number_img = self.image_resize(number_picture[d],column_key)
                        c_x,c_y = self.centering(number_img,column_key)
                        if(j==1):
                            c_x = int(c_x * -1)
                        base_image.paste(number_img,
                                        (column_point["x"]+size["x"]*(j)+c_x,column_point["y"]+size["y"]*i+c_y),
                                        number_img)
                        j += 1    
                elif (column_key == "time"):
                    start_time = val["開始時間"].strftime("%H:%M")
                    j = 0
                    for d in start_time:
                        number_img = self.image_resize(number_picture[d],column_key)
                        c_x,c_y = self.centering(number_img,column_key)
                        if(j==1):
                            c_x = int(c_x * -1)
                        base_image.paste(number_img,
                                        (column_point["x"]+size["x"]*(j)+c_x,column_point["y"]+size["y"]*i+c_y),
                                        number_img)
                        j += 1
                elif(column_key == "content"):
                    c_x,c_y = self.centering(contents_images[val["内容"]],column_key)
                    base_image.paste(contents_images[val["内容"]] ,
                                    (column_point["x"]+c_x,column_point["y"]+size["y"]*i+c_y),
                                    contents_images[val["内容"]])
                elif(column_key == "liver"):
                    j = 0
                    for liver_name,state in val["liver"].items():
                        if(state == 'o'):
                            c_x,c_y = self.centering(liver_images[liver_name],column_key)
                            base_image.paste(liver_images[liver_name],
                                            (column_point["x"]+size["x"]*(j)+c_x,column_point["y"]+size["y"]*i+c_y),
                                            liver_images[liver_name])
                        j += 1
                elif(column_key == "platform"):
                    c_x,c_y = self.centering(platform_images[val["サイト"]],column_key)
                    base_image.paste(platform_images[val["サイト"]] ,
                                    (column_point["x"]+c_x,column_point["y"]+size["y"]*i+c_y))
            i += 1
        """
        for i in range(30):
            draw.line((0, 145+70*i, 1920 ,145+70*i), fill=(255, 0, 255), width=1)
        """
        """
        for key,val in liver_images.items():
            base_image.paste(val,(0,200*i))
            i += 1
        """
        if (ctx == None):
            base_image.show()
        else:
            os.makedirs(self.tmp_dir_path, exist_ok=True)
            img_path = os.path.join(self.tmp_dir_path,"schedule.png")
            base_image.save(img_path)
            await ctx.send(file=discord.File(img_path))

    def get_schedule(self):#エクセルファイルからスケジュールを取得する
        schedule = {}
        label = []
        liver = []
        excel_file_id = self.schedule.split("/")[5]
        excle_wb = self.drive.CreateFile({'id': excel_file_id})
        label_start_idx = 3
        liver_start_idx = 7
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_file_path = '{}/schedule.xlsx'.format(temp_dir)
            excle_wb.GetContentFile(excel_file_path)
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.worksheets[0]
            #labelの取得
            for i in range(4):
                content = ws.cell(2, label_start_idx+i)#C2セルから
                label.append(content.value)
            #ライバーの取得
            i = 0
            while True:
                liver_name = ws.cell(2, liver_start_idx+i)#G2セルから
                if(liver_name.value == None):
                    break
                liver.append(liver_name.value)
                i += 1
            #スケジュールの取得
            for i in range(7):
                date = ws.cell(3+i, 2)
                schedule[str(date.value)] = {"liver":{}}
                #配信情報を取得
                for j in range(len(label)):
                    content = ws.cell(3+i, label_start_idx+j)
                    schedule[str(date.value)][label[j]] = content.value
                for j in range(len(liver)):
                    liver_state = ws.cell(3+i, liver_start_idx+j)
                    schedule[str(date.value)]["liver"][liver[j]] = liver_state.value
            wb.close()
        return schedule

    def image_resize(self,img,column_name):
        img_x,img_y = img.size
        column_x_size = self.column_size[column_name]["x"]
        column_y_size = self.column_size[column_name]["y"]
        content_scale = column_x_size / img_x if  column_x_size / img_x < column_y_size / img_y else column_y_size / img_y
        content_scale = content_scale-0.05
        return img.resize((int(img_x*content_scale),int(img_y*content_scale)))
        

    def get_image(self,img_id):
        img = self.drive.CreateFile({'id': img_id})
        os.makedirs(self.tmp_dir_path, exist_ok=True)
        img_path = os.path.join(self.tmp_dir_path,"{}.png".format(img_id))
        img.GetContentFile(img_path)
        image_obj = Image.open(img_path)
        return image_obj

    def get_schedule_base_image(self):
        return self.get_image(self.config["files"]["format"].split("/")[5])
    
    def get_icons(self,icon_paths,grid_name):
        icon_objects = {}
        for key,val in icon_paths.items():
            if(grid_name == "number_img"):
                icon_objects[key]=self.get_image(val.split("/")[5])
            else:
                icon_objects[key]=self.image_resize(self.get_image(val.split("/")[5]),grid_name)
        return icon_objects

    def get_liver_icon(self):
        return self.get_icons(self.icons,"liver")
        
    def get_live_contents_icon(self):
        return self.get_icons(self.contents,"content")

    def get_live_platform_icon(self):
        return self.get_icons(self.platform,"platform")

    def get_number_picture(self):
        return self.get_icons(self.number_img,"number_img")

if __name__ == "__main__":#検証用コード
    config_file = open("../config/config.json","r",encoding="utf-8")
    config = json.load(config_file)
    function_config = config["function"]
    live_scheduler = LiveScheduer(function_config["live_scheduler"])
    live_scheduler.print_schedule()
    #for key,val in pil_imgs.items():
    #    
    shutil.rmtree("tmp")